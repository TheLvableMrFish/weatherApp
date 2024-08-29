from openpyxl import Workbook

def write_to_excel(city_name, current_weather, high_temps, low_temps):

    current_weather = current_weather
    wb = Workbook()
    ws = wb.active
    ws.title = f"Weather Data for {city_name}"
    ws.append([
        
    ])
    titles = [
        "City", 
        "Current Temperature", 
        "Day 1 Max Temp", "Day 1 Min Temp", 
        "Day 2 Max Temp", "Day 2 Min Temp",
        "Day 3 Max Temp", "Day 3 Min Temp", 
        "Day 4 Max Temp", "Day 4 Min Temp",
        "Day 5 Max Temp", "Day 5 Min Temp", 
        "Day 6 Max Temp", "Day 6 Min Temp",
        "Day 7 Max Temp", "Day 7 Min Temp"
        ]
    
    # Goes through each title in our list of titles and adds them as headers for the excel sheet
    for i, title in enumerate(titles):
        ws.cell(row=1, column= i + 1 , value=title)
    
    # Hard codded the two values cit_name and current_weather
    ws.cell(row=2, column=1, value=city_name)
    ws.cell(row=2, column=2, value=current_weather)
    
    # uses a odd system to put the correct temps inside the excel under the correct name
    j = 0
    for i, temp in enumerate(high_temps):
        ws.cell(row=2, column= 3 + j * 2 , value=temp)
        j += 1

    j = 0
    for i, temp in enumerate(low_temps, 2):
        ws.cell(row=2, column= 4 + j * 2 , value=temp)
        j += 1
    
    # So we can have multiple files with the cityname as there title
    wb.save(f"./{city_name}.xlsx")